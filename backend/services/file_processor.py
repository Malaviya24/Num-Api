import csv
import io
import os
from openpyxl import load_workbook
import psycopg2
from ..config import settings
from ..database import SyncSessionLocal, sync_engine
from ..models import ImportTask
from .number_utils import normalize_mobile

def process_file_sync(task_id: str, filepath: str):
    db = SyncSessionLocal()
    task = db.query(ImportTask).filter(ImportTask.task_id == task_id).first()
    if not task:
        db.close()
        return

    task.status = "PROCESSING"
    db.commit()

    try:
        if filepath.lower().endswith('.csv'):
            total_rows = process_csv(filepath, task_id, db)
        elif filepath.lower().endswith('.xlsx'):
            total_rows = process_xlsx(filepath, task_id, db)
        else:
            raise ValueError("Unsupported file format")

        task.status = "COMPLETED"
        task.total_rows = total_rows
        task.processed_rows = total_rows
        db.commit()
    except Exception as e:
        task.status = "FAILED"
        task.error_message = str(e)
        db.commit()
    finally:
        db.close()

def bulk_insert_copy(rows):
    if not rows: return
    
    conn = sync_engine.raw_connection()
    try:
        with conn.cursor() as cur:
            csv_file = io.StringIO()
            writer = csv.writer(csv_file)
            for row in rows:
                writer.writerow(row)
            csv_file.seek(0)
            
            columns = ('mobile_number', 'full_name', 'email', 'address', 'city', 'state', 'pincode', 'company', 'source_file')
            sql = f"COPY customers ({','.join(columns)}) FROM STDIN WITH CSV"
            cur.copy_expert(sql, csv_file)
        conn.commit()
    finally:
        conn.close()

def update_progress(db, task_id: str, processed: int):
    task = db.query(ImportTask).filter(ImportTask.task_id == task_id).first()
    if task:
        task.processed_rows = processed
        db.commit()

def extract_val(row_dict, keys, max_len=None):
    for k in keys:
        if k in row_dict and row_dict[k]:
            val = str(row_dict[k]).strip()
            if max_len and len(val) > max_len:
                val = val[:max_len]
            return val
    return ''

def process_csv(filepath: str, task_id: str, db) -> int:
    filename = os.path.basename(filepath)
    batch_size = 20000
    batch = []
    total_processed = 0
    
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        reader = csv.DictReader(f)
        for row in reader:
            row_dict = {k.lower().strip() if k else '': v for k, v in row.items()}
            mobile_raw = extract_val(row_dict, ['mobile_number', 'mobile', 'phone', 'phone_number', 'contact'])
            mobile = normalize_mobile(mobile_raw)
            if not mobile: continue
            
            batch.append([
                mobile[:15],
                extract_val(row_dict, ['full_name', 'name', 'customer_name'], 255),
                extract_val(row_dict, ['email', 'email_id'], 255),
                extract_val(row_dict, ['address', 'street']),
                extract_val(row_dict, ['city'], 100),
                extract_val(row_dict, ['state'], 100),
                extract_val(row_dict, ['pincode', 'zip', 'zipcode', 'pin'], 20),
                extract_val(row_dict, ['company', 'organization'], 255),
                filename[:255]
            ])
            
            if len(batch) >= batch_size:
                bulk_insert_copy(batch)
                total_processed += len(batch)
                batch = []
                update_progress(db, task_id, total_processed)
                
        if batch:
            bulk_insert_copy(batch)
            total_processed += len(batch)
            update_progress(db, task_id, total_processed)
            
    return total_processed

def process_xlsx(filepath: str, task_id: str, db) -> int:
    filename = os.path.basename(filepath)
    batch_size = 20000
    batch = []
    total_processed = 0
    
    wb = load_workbook(filename=filepath, read_only=True, data_only=True)
    ws = wb.active
    
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).lower().strip() if h else '' for h in row]
            continue
            
        row_dict = dict(zip(headers, row))
        mobile_raw = extract_val(row_dict, ['mobile_number', 'mobile', 'phone', 'phone_number', 'contact'])
        mobile = normalize_mobile(mobile_raw)
        if not mobile or mobile == 'None': continue
        
        batch.append([
            mobile[:15],
            extract_val(row_dict, ['full_name', 'name', 'customer_name'], 255),
            extract_val(row_dict, ['email', 'email_id'], 255),
            extract_val(row_dict, ['address', 'street']),
            extract_val(row_dict, ['city'], 100),
            extract_val(row_dict, ['state'], 100),
            extract_val(row_dict, ['pincode', 'zip', 'zipcode', 'pin'], 20),
            extract_val(row_dict, ['company', 'organization'], 255),
            filename[:255]
        ])
        
        if len(batch) >= batch_size:
            bulk_insert_copy(batch)
            total_processed += len(batch)
            batch = []
            update_progress(db, task_id, total_processed)
            
    if batch:
        bulk_insert_copy(batch)
        total_processed += len(batch)
        update_progress(db, task_id, total_processed)
        
    wb.close()
    return total_processed
